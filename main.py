from db.models import Record
import openpyxl
# import logging
# logger = logging.getLogger(__name__)


from openpyxl import load_workbook

import os

excel_files = os.listdir('files')

Record.objects.all().delete()

for file_name in excel_files[:1]:
    wb = load_workbook(filename = 'files/'+file_name)
    worksheet = wb.worksheets[0]
    header_list = []
    for row in worksheet.iter_rows():
        # logger.debug(row)
        if not header_list:
            for cell in row:
                header_list.append(cell.value)
            print("header_list =", header_list)
            ac_reg_idx = header_list.index("A/C REG")
            type_idx = header_list.index("TYPE")
            arr_idx = header_list.index("ARR")
            dep_idx = header_list.index("DEP")
            route_idx = header_list.index("ROUTE")
            date_idx = header_list.index("DATE")
            eta_time_idx = header_list.index("ETA")
            etd_time_idx = header_list.index("ETD")

        else:
            print("============================================================")
            o = Record()
            o.ac_reg = str(row[ac_reg_idx].value)
            o.type = str(row[type_idx].value)
            o.arr = str(row[arr_idx].value)
            o.dep = str(row[dep_idx].value)
            # o.date = str(row[date_idx].value)
            o.route = str(row[route_idx].value)
            o.time_eta = str(row[eta_time_idx].value)

            if ":" in o.time_eta:
                eta_list = o.time_eta.split(":")
                print(eta_list)
                eta_hour = eta_list[0]
                eta_min = eta_list[1]

                if "+" in eta_min:
                    o.day_change = 1
                    print("day change")
                    eta_min = eta_min.replace("+","")
                print("eta_hour = %s" %(eta_hour))
                print("eta_min = %s" %(eta_min))

                o.time_eta = str(eta_hour)+":"+str(eta_min)


            o.time_etd = str(row[etd_time_idx].value)
            print("o.ac_reg = %s" %(o.ac_reg))
            print("o.type = %s" %(o.type))
            print("o.arr = %s" %(o.arr))
            print("o.date = %s" %(str(row[date_idx].value)))

            print("o.date = %s" %(worksheet["M2"].value))

            print("o.dep = %s" %(o.dep))
            print("o.route = %s" %(o.route))
            print("o.time_eta = %s" %(o.time_eta))
            print("o.time_etd = %s" %(o.time_etd))
            o.save()

# # data = Record.objects.filter(type="A350", route__icontains="SFO")
# # print("==============================================================")
# # print("==============================================================")
# # print("==============================================================")
# # print("==============================================================")
# # for o in data:
# #     print("+++++++++++++++++++++++++++++++++++++++++++++++++++++")
# #     print("o.ac_reg = %s" % (o.ac_reg))
# #     print("o.type = %s" % (o.type))
# #     print("o.arr = %s" % (o.arr))
# #     print("o.date = %s" % (str(row[date_idx].value)))
# #
# #     print("o.date = %s" % (worksheet["M2"].value))
# #
# #     print("o.dep = %s" % (o.dep))
# #     print("o.route = %s" % (o.route))
# #     print("o.eta = %s" % (o.eta))
# #     print("o.etd = %s" % (o.etd))
#
#
# print("++++++++++++++++++++++++++++++++++++++++++++++++++")
# print("++++++++++++++++++++++++++++++++++++++++++++++++++")
# print("++++++++++++++++++++++++++++++++++++++++++++++++++")
#
data = Record.objects.all()
print("data.count = %s" %(data.count()))



type_list = data.order_by('type').values_list('type', flat=True).distinct()
print("type_list = %s" %(type_list))
for type in type_list:
    print("type = %s count = %s" %(type, data.filter(type=type).count()))

route_list = data.order_by('route').values_list('route', flat=True).distinct()
print("route_list = %s" %(route_list))
print("len(route_list) = %s" %(len(route_list)))
for route in route_list:
    print("route = %s count = %s" %(route, data.filter(route=route).count()))




from openpyxl import Workbook
wb = Workbook()
# wb = load_workbook ('data.xlsx')
# ws1 = wb.sheetnames()
wb.create_sheet("sheet1")
ws1 = wb["sheet1"]


offset_row = 5
offset_col = 0

row = 1
for o in data:
    col = 0

    ws1.cell(row, col+1, o.ac_reg)
    ws1.cell(row, col+2, o.type)
    ws1.cell(row, col+3, o.arr)
    ws1.cell(row, col+4, o.dep)
    ws1.cell(row, col+5, o.route)
    ws1.cell(row, col+6, o.time_eta)
    ws1.cell(row, col+7, o.time_etd)

    row += 1


os.remove('data.xlsx')

wb.save('data.xlsx')
